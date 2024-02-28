from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
import logging


def create_excel_report(df):
    """
    Create an Excel report from a dataframe with the following columns: 'Channel', 'Status', 'Next estimated free slot'.
    :param df: Dataframe with the following columns: 'Channel', 'Status', 'Next estimated free slot'.
    :return: Workbook object
    """
    # Log the creation of the Excel report
    logging.debug(f"Creating Excel report from dataframe")

    # Sort the dataframe by 'Channel' alphabetically
    df = df.sort_values(by='Channel')

    try:
        df.drop(['Next estimated free slot', 'Year', 'Availability'], axis=1, inplace=True)
    except KeyError:
        pass

    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Write the dataframe to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Define the font
    font = Font(name='Northvolt Grit')

    # Define the fill colors
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Apply the font and fill colors
    for row in ws.iter_rows():
        for cell in row:
            cell.font = font
            if cell.value == 'Empty':
                cell.fill = green_fill
            elif cell.value == 'Active':
                cell.fill = red_fill
    logging.debug(f"Excel report created")
    return wb


def save_excel_report(wb, filename):
    """
    Save the workbook to a file
    :param wb: Workbook object
    :param filename: File name to save the workbook
    :return: None
    """
    # Log the saving of the workbook
    logging.debug(f"Saving Excel report to {filename}")

    # Save the workbook
    wb.save(filename)

    # Log the end of the function
    logging.debug(f"Excel report saved to {filename}")